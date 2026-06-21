import shutil
import tempfile
from io import BytesIO

import pandas as pd
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import override_settings
from openpyxl import Workbook, load_workbook
from rest_framework.test import APITestCase

from datasets.models import Dataset, TransformRun


class TransformationApplyApiTests(APITestCase):
    def setUp(self):
        self.media_root = tempfile.mkdtemp()
        self.settings_override = override_settings(MEDIA_ROOT=self.media_root)
        self.settings_override.enable()
        upload = SimpleUploadedFile(
            "customers.csv",
            (
                b"name,email,notes\n"
                b"Ada,ada@example.com,Contact ada@example.com\n"
                b"Lin,lin@example.com,=2+2\n"
            ),
            content_type="text/csv",
        )
        response = self.client.post("/api/datasets/", {"file": upload}, format="multipart")
        self.dataset = Dataset.objects.get(id=response.data["id"])
        self.url = f"/api/datasets/{self.dataset.id}/transforms/apply/"
        self.payload = {
            "instruction": "Find email addresses",
            "pattern": r"\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}\b",
            "replacement": "[REDACTED]",
            "columns": ["email", "notes"],
            "flags": [],
            "explanation": "Matches email addresses.",
            "provider": "built-in",
            "model": "common-patterns-v1",
        }

    def tearDown(self):
        self.settings_override.disable()
        shutil.rmtree(self.media_root, ignore_errors=True)

    def test_applies_transform_and_persists_auditable_run(self):
        response = self.client.post(self.url, self.payload, format="json")

        self.assertEqual(response.status_code, 201)
        self.assertEqual(response.data["match_count"], 3)
        self.assertEqual(response.data["affected_rows"], 2)
        self.assertEqual(response.data["provider"], "built-in")
        self.assertEqual(TransformRun.objects.count(), 1)
        run = TransformRun.objects.get()
        self.assertTrue(run.result_file.name.endswith(".csv"))
        self.assertEqual(run.dataset, self.dataset)

    def test_downloads_processed_csv_without_mutating_source(self):
        original_bytes = self.dataset.source_file.read()
        created = self.client.post(self.url, self.payload, format="json")

        response = self.client.get(created.data["download_url"])
        downloaded = b"".join(response.streaming_content)
        frame = pd.read_csv(BytesIO(downloaded))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(frame.loc[0, "email"], "[REDACTED]")
        self.assertEqual(frame.loc[0, "notes"], "Contact [REDACTED]")
        self.assertEqual(frame.loc[1, "notes"], "'=2+2")
        self.dataset.source_file.open("rb")
        self.assertEqual(self.dataset.source_file.read(), original_bytes)

    def test_revalidates_pattern_before_apply(self):
        self.payload["pattern"] = "(a+)+$"

        response = self.client.post(self.url, self.payload, format="json")

        self.assertEqual(response.status_code, 400)
        self.assertEqual(response.data["code"], "unsafe_transform")
        self.assertEqual(TransformRun.objects.count(), 0)

    def test_applies_and_downloads_excel_in_its_original_format(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Customers"
        sheet.append(["notes"])
        sheet.append(["Email ada@example.com"])
        source = BytesIO()
        workbook.save(source)
        upload = SimpleUploadedFile(
            "customers.xlsx",
            source.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        created_dataset = self.client.post(
            "/api/datasets/",
            {"file": upload},
            format="multipart",
        )
        url = f"/api/datasets/{created_dataset.data['id']}/transforms/apply/"
        payload = {**self.payload, "columns": ["notes"]}

        created_run = self.client.post(url, payload, format="json")
        response = self.client.get(created_run.data["download_url"])
        processed = load_workbook(BytesIO(b"".join(response.streaming_content)))

        self.assertEqual(created_run.status_code, 201)
        self.assertEqual(created_run.data["output_format"], "xlsx")
        self.assertEqual(processed["Customers"]["A2"].value, "Email [REDACTED]")
