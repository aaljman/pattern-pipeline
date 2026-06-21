import shutil
import tempfile
from datetime import time
from io import BytesIO
from unittest.mock import patch

from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import override_settings
from openpyxl import Workbook
from rest_framework.test import APITestCase

from datasets.models import Dataset


class DatasetApiTests(APITestCase):
    def setUp(self):
        self.media_root = tempfile.mkdtemp()
        self.settings_override = override_settings(MEDIA_ROOT=self.media_root)
        self.settings_override.enable()

    def tearDown(self):
        self.settings_override.disable()
        shutil.rmtree(self.media_root, ignore_errors=True)

    def test_uploads_and_profiles_csv(self):
        upload = SimpleUploadedFile(
            "customers.csv",
            b"name,email,score\nAda,ada@example.com,10\nLin,,20\n",
            content_type="text/csv",
        )

        response = self.client.post("/api/datasets/", {"file": upload}, format="multipart")

        self.assertEqual(response.status_code, 201)
        self.assertEqual(response.data["row_count"], 2)
        self.assertEqual(response.data["text_columns"], ["name", "email"])
        self.assertEqual(response.data["columns"][1]["missing_count"], 1)
        self.assertIsNone(response.data["preview"][1]["email"])
        self.assertEqual(Dataset.objects.count(), 1)

    def test_uploads_first_sheet_from_xlsx(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Customers"
        sheet.append(["name", "phone"])
        sheet.append(["Ada", "0412 345 678"])
        buffer = BytesIO()
        workbook.save(buffer)
        upload = SimpleUploadedFile(
            "customers.xlsx",
            buffer.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        response = self.client.post("/api/datasets/", {"file": upload}, format="multipart")

        self.assertEqual(response.status_code, 201)
        self.assertEqual(response.data["sheet_name"], "Customers")
        self.assertEqual(response.data["preview"][0]["phone"], "0412 345 678")

    def test_normalizes_numeric_headers_and_time_values_from_xlsx(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.append([123, "opens_at"])
        sheet.append(["value", time(9, 30)])
        buffer = BytesIO()
        workbook.save(buffer)
        upload = SimpleUploadedFile(
            "mixed.xlsx",
            buffer.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        response = self.client.post("/api/datasets/", {"file": upload}, format="multipart")

        self.assertEqual(response.status_code, 201)
        self.assertEqual(response.data["columns"][0]["name"], "123")
        self.assertEqual(response.data["preview"][0]["opens_at"], "09:30:00")

    def test_rejects_fake_xlsx_zip_archive(self):
        upload = SimpleUploadedFile(
            "fake.xlsx",
            b"PK\x03\x04not-a-real-archive",
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        response = self.client.post("/api/datasets/", {"file": upload}, format="multipart")

        self.assertEqual(response.status_code, 400)
        self.assertEqual(response.data["code"], "invalid_file")

    def test_rejects_unsupported_file_type(self):
        upload = SimpleUploadedFile("customers.txt", b"name\nAda\n", content_type="text/plain")

        response = self.client.post("/api/datasets/", {"file": upload}, format="multipart")

        self.assertEqual(response.status_code, 400)
        self.assertEqual(response.data["code"], "invalid_file")
        self.assertEqual(Dataset.objects.count(), 0)

    def test_rejects_empty_files(self):
        upload = SimpleUploadedFile("empty.csv", b"", content_type="text/csv")

        response = self.client.post("/api/datasets/", {"file": upload}, format="multipart")

        self.assertEqual(response.status_code, 400)
        self.assertIn("empty", response.data["message"])

    @patch("datasets.services.ingestion.MAX_ROWS", 1)
    def test_rejects_files_over_row_limit(self):
        upload = SimpleUploadedFile(
            "rows.csv",
            b"name\nAda\nLin\n",
            content_type="text/csv",
        )

        response = self.client.post("/api/datasets/", {"file": upload}, format="multipart")

        self.assertEqual(response.status_code, 400)
        self.assertIn("at most 1 rows", response.data["message"])

    @patch("datasets.services.ingestion.MAX_COLUMNS", 2)
    def test_rejects_files_over_column_limit(self):
        upload = SimpleUploadedFile(
            "columns.csv",
            b"one,two,three\n1,2,3\n",
            content_type="text/csv",
        )

        response = self.client.post("/api/datasets/", {"file": upload}, format="multipart")

        self.assertEqual(response.status_code, 400)
        self.assertIn("at most 2 columns", response.data["message"])

    @patch("datasets.services.ingestion.MAX_XLSX_UNCOMPRESSED_BYTES", 1)
    def test_rejects_xlsx_over_expanded_size_limit(self):
        workbook = Workbook()
        workbook.active.append(["name"])
        workbook.active.append(["Ada"])
        buffer = BytesIO()
        workbook.save(buffer)
        upload = SimpleUploadedFile(
            "expanded.xlsx",
            buffer.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        response = self.client.post("/api/datasets/", {"file": upload}, format="multipart")

        self.assertEqual(response.status_code, 400)
        self.assertIn("expanded XLSX", response.data["message"])

    def test_returns_dataset_profile(self):
        upload = SimpleUploadedFile("names.csv", b"name\nAda\n", content_type="text/csv")
        created = self.client.post("/api/datasets/", {"file": upload}, format="multipart")

        response = self.client.get(f"/api/datasets/{created.data['id']}/")

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data["original_name"], "names.csv")

    def test_returns_consistent_not_found_error(self):
        response = self.client.get("/api/datasets/00000000-0000-0000-0000-000000000000/")

        self.assertEqual(response.status_code, 404)
        self.assertEqual(response.data["code"], "not_found")
